import os
import sys
import re
import datetime
import subprocess
import openpyxl


#/* 正規表現 */
re_total_lines     = re.compile(r"^Total Lines\s+(\d+)\n")
re_total_steps     = re.compile(r"^Total Steps\s+(\d+)\n")
re_total_comments  = re.compile(r"^Total Comment\s+(\d+)\n")
re_inc_files       = re.compile(r"^Include Files\n")
re_macro_defs      = re.compile(r"^macro defs\n")
re_type_defs       = re.compile(r"^type defs\n")
re_variables       = re.compile(r"^Variables List\n")
re_functions       = re.compile(r"^Function List\n")
re_function_detail = re.compile(r"^Function Detail\n")

re_include_file     = re.compile(r"^\t([^\t]+)\n$")
re_variables_line   = re.compile(r"^\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\n$")
re_functions_line   = re.compile(r"^\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\t([^\t]+)\n$")
re_detail_name      = re.compile(r"^\tName  \t([^\n]+)\n$")
re_detail_local     = re.compile(r"^\tLocal Variable\n$")
re_detail_func2call = re.compile(r"^\tFunctions call to\n$")
re_detail_called    = re.compile(r"^\tcalled from\n$")
re_detail_var_wr    = re.compile(r"^\tVariables write to\n$")
re_detail_var_rd    = re.compile(r"^\tVariables read from\n$")
re_detail_array     = re.compile(r"^\t\[(\d+)\]\t([^\n]+)\n$")

re_array_variable   = re.compile(r"([_A-Za-z][_A-Za-z0-9]*)\[.*\]")

re_typedef          = re.compile(r"^\t([^\t]+)\t(struct|union|enum)[^\t]+\t([^\n]+)")
re_typedef_normal   = re.compile(r"^\t([^\t]+)\t([^\t]+)\t([^\n]+)")
re_typedef_member   = re.compile(r"^\t\t([^\t]+)\t([^\t]*)\t([^\t]+)\t([^\t]+)\n")


#/* グローバル変数 */
g_output_fld         = "m_analyze"
g_setting_file       = "m_analyze_setting.txt"
g_c_setting_file    = "c_analyze_setting.txt"
g_log_file_name      = ""
g_output_temp_text   = 0
g_default_log        = 0
g_target_files       = []
g_module_name        = "sample"
g_jar_path           = ""
g_exec_c_analyze     = 1
g_charset_utf        = 0
g_without_pu_convert = 0
g_modules            = []
g_call_tree          = []
g_called_tree        = []
g_tree_stack         = []
g_tree_depend        = []


class cAnalyzeSettings:
    def __init__(self, file_name, setting_file, char_set):
        self.target_file  = file_name
        self.setting_file = setting_file
        self.char_set     = char_set
        return


class cDependency:
    def __init__(self):
        self.function    = ""
        self.stack       = []
        return

class cMember:
    def __init__(self):
        self.name        = ""
        self.type        = ""
        self.summary     = ""

class cTypedef:
    def __init__(self):
        self.name        = ""
        self.type        = ""
        self.summary     = ""
        self.members     = []

class cVariable:
    def __init__(self):
        self.name        = ""
        self.type        = ""
        self.summary     = ""
        self.init_val    = ""
        self.is_static   = 0
        self.is_const    = 0
        self.is_extern   = 0
        self.func_read   = []
        self.func_write  = []

class cFunction:
    def __init__(self):
        self.name        = ""
        self.ret_type    = ""
        self.summary     = ""
        self.is_static   = 0
        self.lines       = 0
        self.steps       = 0
        self.comments    = 0
        self.paths       = 0
        self.args        = []
        self.local_vars  = []
        self.func_call   = []
        self.func_called = []
        self.var_read    = []
        self.var_write   = []

class cModule:
    def __init__(self):
        self.name         = ""
        self.lines        = 0
        self.steps        = 0
        self.comments     = 0
        self.paths        = 0
        self.current_mode = ""
        self.includes     = []
        self.macros       = []
        self.typedefs     = []
        self.variables    = []
        self.functions    = []
        self.module_ref   = []
        self.module_refed = []

        #/* 内部処理用変数 */
        self.skip_line    = 0
        self.detail_phase = 0
        self.current_func = 0
        self.current_typedef = None

    def get_function(self, name):
        for function in self.functions:
            if (function.name == name):
                return function

        return


    def get_variable(self, name):
        for variable in self.variables:
            if (variable.name == name):
                return variable

        return


    def check_top_level(self, line):
#       print ("chack_top_level")
        if (result := re_total_lines.match(line)):
            print ("Total Lines  : " + result.group(1))
            self.lines = int(result.group(1))
        elif (result := re_total_steps.match(line)):
            print ("Total Steps  : " + result.group(1))
            self.steps = int(result.group(1))
        elif (result := re_total_comments.match(line)):
            print ("Total Comments  : " + result.group(1))
            self.comments = int(result.group(1))
        elif (result := re_inc_files.match(line)):
            return "include"
        elif (result := re_macro_defs.match(line)):
            self.skip_line = 1
            return "macro"
        elif (result := re_type_defs.match(line)):
            self.skip_line = 1
            return "type"
        elif (result := re_variables.match(line)):
            self.skip_line = 1
            return "variable"
        elif (result := re_functions.match(line)):
            self.skip_line = 1
            return "function"
        elif (result := re_function_detail.match(line)):
            return "detail"

        return "none"

    def parse_include_line(self, line):
#       print("include line : %s" % line)
        if (result := re_include_file.match(line)):
            print("include : %s" % result.group(1))
            self.includes.append(result.group(1))
        return

    def parse_macro_line(self, line):
        return

    def parse_type_line(self, line):
        if (result := re_typedef.match(line)):
            aTypedef = cTypedef()
            aTypedef.name    = result.group(1)
            aTypedef.type    = result.group(2)
            aTypedef.summary = result.group(3)
            print("typedef %s line %s - %s" % (aTypedef.type, aTypedef.name, aTypedef.summary))
            self.current_typedef = aTypedef
            self.typedefs.append(aTypedef)

        elif (result := re_typedef_normal.match(line)):
            aTypedef         = cTypedef()
            aTypedef.name    = result.group(1)
            aTypedef.type    = result.group(2)
            aTypedef.summary = result.group(3)
            print("typedef none line %s - %s - %s" % (aTypedef.type, aTypedef.name, aTypedef.summary))
            self.typedefs.append(aTypedef)
        elif (result := re_typedef_member.match(line)):
            if (self.current_typedef == None):
                print("strange typedef member!!!!")
            else:
                aMember = cMember()
                aMember.summary = result.group(3)
                if (self.current_typedef.type == "enum"):
                    aMember.name = result.group(1)
                    aMember.type = result.group(2)
                    print("enum member %s - %s - %s" % (aMember.name, aMember.type, aMember.summary))
                else:
                    aMember.name = result.group(2)
                    aMember.type = result.group(1)
                    print("struct member %s - %s - %s" % (aMember.type, aMember.name, aMember.summary))

                self.current_typedef.members.append(aMember)

        return

    #/*****************************************************************************/
    #/* 変数概要の読み込み                                                        */
    #/*****************************************************************************/
    def parse_variable_line(self, line):
        if (result := re_variables_line.match(line)):
            print("variable : %s" % result.group(2))
            aVariable = cVariable()
            aVariable.type      = result.group(1)
            aVariable.name      = result.group(2)
            aVariable.init_val  = result.group(3)
            aVariable.summary   = result.group(4)
            aVariable.is_static = int(result.group(5))
            aVariable.is_extern = int(result.group(6))
            return aVariable
        return None

    #/*****************************************************************************/
    #/* 関数概要の読み込み                                                        */
    #/*****************************************************************************/
    def parse_function_line(self, line):
        if (result := re_functions_line.match(line)):
            print("function : %s" % result.group(1))
            aFunction = cFunction()
            aFunction.name      = result.group(1)
            aFunction.ret_type  = result.group(2)
            aFunction.summary   = result.group(3)
            aFunction.is_static = int(result.group(4))
            aFunction.lines = int(result.group(5))
            aFunction.steps = int(result.group(6))
            aFunction.comments = int(result.group(7))
            aFunction.paths = int(result.group(8))
            return aFunction
        return None

    #/*****************************************************************************/
    #/* 関数詳細の読み込み                                                        */
    #/*****************************************************************************/
    def parse_detail_line(self, line):
        if (self.detail_phase == 0):
            #/* まず関数名を拾い、ローカル変数定義の先頭を見つける */
            if (result := re_detail_name.match(line)):
                print("detail : %s" % result.group(1))
                self.current_func = self.get_function(result.group(1))
            elif (result := re_detail_local.match(line)):
                self.detail_phase = 1
                self.skip_line = 1
        elif (self.detail_phase == 1):
            #/* 次に呼び出す関数の先頭を見つける */
            if (result := re_detail_func2call.match(line)):
                self.detail_phase = 2
            else:
                ret_val = self.parse_variable_line(line)
                if (ret_val != None):
                    self.current_func.local_vars.append(ret_val)
        elif (self.detail_phase == 2):
            #/* 次に呼び出されている関数の先頭を見つける */
            if (result := re_detail_called.match(line)):
                self.detail_phase = 3
            elif (result := re_detail_array.match(line)):
                self.current_func.func_call.append(result.group(2))
        elif (self.detail_phase == 3):
            #/* 次に書き込む変数の先頭を見つける */
            if (result := re_detail_var_wr.match(line)):
                self.detail_phase = 4
            elif (result := re_detail_array.match(line)):
                self.current_func.func_called.append(result.group(2))
        elif (self.detail_phase == 4):
            #/* 次に読み出す変数の先頭を見つける */
            if (result := re_detail_var_rd.match(line)):
                self.detail_phase = 5
            elif (result := re_detail_array.match(line)):
                self.current_func.var_write.append(result.group(2))
        elif (self.detail_phase == 5):
            #/* 最後、読み出す変数がヒットしなくなったら、次の関数を探す */
            if (result := re_detail_array.match(line)):
                self.current_func.var_read.append(result.group(2))
            else:
                self.detail_phase = 0
        return


    #/*****************************************************************************/
    #/* コード解析結果の読み込み                                                  */
    #/*****************************************************************************/
    def load_csv(self, file_path):
        global g_charset_utf

        if (g_charset_utf):
            csv = open(file_path, 'r', encoding="utf-8")
        else:
            csv = open(file_path, 'r')

        read_lines = csv.readlines()
        self.skip_line = 0
        for line in read_lines:
            ret_text = self.check_top_level(line)
            if (ret_text != "none"):
                print("ret_text : %s" % ret_text)
                self.current_mode = ret_text
                continue
            else:
                if (self.skip_line == 1):
                    self.skip_line = 0
                    continue

            if (self.current_mode == "include"):
                self.parse_include_line(line)
            elif (self.current_mode == "macro"):
                self.parse_macro_line(line)
            elif (self.current_mode == "type"):
                self.parse_type_line(line)
            elif (self.current_mode == "variable"):
                ret_val = self.parse_variable_line(line)
                if (ret_val != None):
                    print("add variable! : %s" % ret_val.name)
                    self.variables.append(ret_val)
            elif (self.current_mode == "function"):
                ret_val = self.parse_function_line(line)
                if (ret_val != None):
#                   print("add function!")
                    self.functions.append(ret_val)
            elif (self.current_mode == "detail"):
                self.parse_detail_line(line)


        csv.close();


    #/*****************************************************************************/
    #/* モジュール外からの関数コールを拾う                                        */
    #/*****************************************************************************/
    def check_called_func(self):
        global g_modules
        for function in self.functions:
            if (function.is_static):
                break

            for other_module in g_modules:
                if (other_module.name == self.name):
                    continue

                for other_func in other_module.functions:
                    for call in other_func.func_call:
                        if (call == function.name):
                            function.func_called.append(other_func.name)



    #/*****************************************************************************/
    #/* スタティック変数のチェック                                                */
    #/*****************************************************************************/
    def check_static_variable(self):
#       print("check_static_variable! : %s" % self.name)
#       print("lines : %d" % self.lines)
        for variable in self.variables:
            if (variable.is_static):
                print("static variable : %s" % variable.name)
                tmp_name = variable.name
                if (result := re_array_variable.match(tmp_name)):
#                   print("match array!")
                    tmp_name = result.group(1)

                for function in self.functions:
                    for var in function.var_read:
                        if (var == tmp_name):
                            print("  read by    : %s" % function.name)
                            variable.func_read.append(function.name)

                for function in self.functions:
                    for var in function.var_write:
                        if (var == tmp_name):
                            print("  written by : %s" % function.name)
                            variable.func_write.append(function.name)

#/*****************************************************************************/
#/* サブディレクトリの生成                                                    */
#/*****************************************************************************/
def make_directory(dirname):
    print("make dir! %s" % dirname)
    os.makedirs(os.path.join(dirname), exist_ok = True)



def search_dir(directory):
    global g_target_files

#   print ("search_dir %s" % directory)
    dirlist = []
    files = os.listdir(directory)
    for filename in files:
#       print("filename : %s" % filename)
        if (os.path.isdir(directory + "\\" + filename)):
            search_dir(directory + "\\" + filename)
        else:
            result = re.match(r".*\.[cChH]$", filename)
            if (result):
#               print ("Match! filename : %s" % filename)
#               g_target_files.append(directory + "\\" + filename)
                c_analyze_set = cAnalyzeSettings(directory + "\\" + filename, "", 0)
                g_target_files.append(c_analyze_set)


def check_command_line_option():
    global g_setting_file
    global g_c_setting_file
    global g_output_fld
    global g_log_file_name
    global g_default_log
    global g_target_files
    global g_output_temp_text
    global g_exec_c_analyze
    global g_charset_utf
    global g_without_pu_convert

    argc = len(sys.argv)
    option = ""
    first  = 1

    for arg in sys.argv:
        if (first == 1):
            first = 0
            continue

#       print (arg)
        if (option == "s"):
            option = ""
            g_setting_file = arg
        elif (option == "o"):
            option = ""
            g_output_fld = arg
        elif (option == "l"):
            option = ""
            g_log_file_name = arg
        elif (arg == "-s"):
            option = "s"
        elif (arg == "-l"):
            option = "l"
        elif (arg == "-skip"):
            g_exec_c_analyze = 0
        elif (arg == "-utf"):
            g_charset_utf = 1
        elif (arg == "-nopu"):
            g_without_pu_convert = 1
        elif (arg == "-o"):
            option = "o"
        elif (arg == "-t"):
            g_output_temp_text = 1
        elif (arg == "-dl"):
            print("default log file!")
            g_default_log = 1
        else:
            if (os.path.isdir(arg)):
#               print("directory! %s" % arg)
                search_dir(arg)
            elif (os.path.isfile(arg)):
#               print("file! %s" % arg)
#               g_target_files.append(arg)
                c_analyze_set = cAnalyzeSettings(arg, "", 0)
                g_target_files.append(c_analyze_set)
            else:
                print("unknown arg! %s" % arg)


    if (os.path.isfile(g_setting_file) == False):
        print("setting file not found! [%s]" % g_setting_file)
        print("  specify settning file with -s option.")
        exit(-1)

    if (argc == 1):
        print("analyze module by default setting.txt")



#/*****************************************************************************/
#/* 設定ファイルの読み込み処理                                                */
#/*****************************************************************************/
def read_setting_file():
    global g_target_files
    global g_charset_utf
    global g_module_name
    global g_call_tree
    global g_called_tree
    global g_jar_path
    global g_c_setting_file

    print("read_setting_file");
    f = open(g_setting_file, 'r')
    lines = f.readlines()

    re_source     = re.compile(r"source\s+(.+)\n")
    re_jar        = re.compile(r"plantuml[ \t]+([^\s]+)")
    re_c_setting  = re.compile(r"c_setting([0-9]*)[ \t]+([^\s]+)")
    re_module     = re.compile(r"module_name[ \t]+([^\s]+)")
    re_charset    = re.compile(r"default_charset[ \t]+([^\s]+)")
    re_calltree   = re.compile(r"calltree[ \t]+([^\s]+)")
    re_calledtree = re.compile(r"calledtree[ \t]+([^\s]+)")

    for line in lines:
#       print ("line:%s" % line)
        if (result := re_source.match(line)):
            print ("source   : " + result.group(1))
#           g_target_files.append(result.group(1))
            c_analyze_set = cAnalyzeSettings(result.group(1), g_c_setting_file, g_charset_utf)
            g_target_files.append(c_analyze_set)
        elif (result := re_c_setting.match(line)):
            g_c_setting_file = result.group(2)
            print ("c_setting file : " + g_c_setting_file)
        elif (result := re_jar.match(line)):
            print ("jar file : " + result.group(1))
            g_jar_path = result.group(1)
        elif (result := re_module.match(line)):
            print ("module   : " + result.group(1))
            g_module_name = result.group(1)
        elif (result := re_calltree.match(line)):
            print ("calltree : " + result.group(1))
            g_call_tree.append(result.group(1))
        elif (result := re_calledtree.match(line)):
            print ("calledtree : " + result.group(1))
            g_called_tree.append(result.group(1))
        elif (result := re_charset.match(line)):
            print ("charset  : " + result.group(1))
            if (result.group(1) == "UTF8"):
                g_charset_utf = 1
            else:
                g_charset_utf = 0

    f.close()


#/*****************************************************************************/
#/* ログファイル設定                                                          */
#/*****************************************************************************/
def log_settings():
    global g_log_file_name
    global g_output_fld

    log_path = ""

    print("log_settings")
    if (g_log_file_name != ""):
        log_path = g_output_fld + "\\" + g_log_file_name
    elif (g_default_log == 1):
        now = datetime.datetime.now()
        formatted_time = now.strftime("%Y%m%d_%H%M%S")
        log_path = g_output_fld + "\\m_analyze_log_" + formatted_time + ".log";

    print ("log_path : %s" % log_path)

    if (log_path != ""):
       log_file = open(log_path, "a")
       sys.stdout = log_file



#/*****************************************************************************/
#/* Cコード解析                                                               */
#/*****************************************************************************/
def c_analyze():
    global g_target_files
    global g_without_pu_convert

    print("c_analyze")
    for source_file in g_target_files:
        print("  analyzing %s" % source_file.target_file, file = sys.__stdout__)

        cmd_text = "perl c_analyze.pl -o %s %s" % (g_output_fld, source_file.target_file)

        if (source_file.setting_file != ""):
            cmd_text += " -s " + source_file.setting_file

        if (g_output_temp_text == 1):
            cmd_text += " -t"

        if (source_file.char_set == 1):
            cmd_text += " -utf"

        if (g_without_pu_convert == 1):
            cmd_text += " -nopu"

        subprocess.run(cmd_text, stdout=sys.stdout)

#/*****************************************************************************/
#/* モジュール解析                                                            */
#/*****************************************************************************/
def module_analyze():
    global g_target_files
    global g_modules

    print("module_analyze")
    for source_file in g_target_files:
        csv_file_name = os.path.basename(source_file.target_file) + "_analyzed.csv"
        print("read csv : %s" % csv_file_name, file = sys.__stdout__);
        csv_file_name = g_output_fld + "\\" + csv_file_name
        module = cModule()
        module.name = os.path.basename(source_file.target_file)
        module.load_csv(csv_file_name)
        g_modules.append(module)

    for module in g_modules:
        module.check_static_variable()

    for module in g_modules:
        module.check_called_func()



#/*****************************************************************************/
#/* 変数一覧出力                                                              */
#/*****************************************************************************/
def out_variable_list():
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "変数一覧"

    row = 2
    col = 2
    ws.cell(row, col    ).value = "ファイル名"
    ws.cell(row, col + 1).value = "型"
    ws.cell(row, col + 2).value = "変数名"
    ws.cell(row, col + 3).value = "説明"

    row += 1
    for module in g_modules:
        for variable in module.variables:
            ws.cell(row, col    ).value = module.name
            ws.cell(row, col + 1).value = variable.type
            ws.cell(row, col + 2).value = variable.name
            ws.cell(row, col + 3).value = variable.summary
            row += 1


    ws = wb.create_sheet("変数一覧(詳細)")
    row = 2
    col = 2
    ws.cell(row, col    ).value = "ファイル名"
    ws.cell(row, col + 1).value = "型"
    ws.cell(row, col + 2).value = "変数名"
    ws.cell(row, col + 3).value = "説明"
    ws.cell(row, col + 4).value = "参照している関数"
    ws.cell(row, col + 5).value = "更新している関数"
    row += 1
    for module in g_modules:
        for variable in module.variables:
            ws.cell(row, col    ).value = module.name
            ws.cell(row, col + 1).value = variable.type
            ws.cell(row, col + 2).value = variable.name
            ws.cell(row, col + 3).value = variable.summary

            row_start = row
            for reader in variable.func_read:
                ws.cell(row_start, col + 4).value = reader
                row_start += 1
            row_end = row_start

            row_start = row
            for writer in variable.func_write:
                ws.cell(row_start, col + 5).value = writer
                row_start += 1

            if (row_start > row_end):
                row_end = row_start

            if (row_end > row + 1):
                #/* セル結合処理 */
                ws.merge_cells(start_row = row, end_row = row_end - 1, start_column = 2, end_column = 2)
                ws.merge_cells(start_row = row, end_row = row_end - 1, start_column = 3, end_column = 3)
                ws.merge_cells(start_row = row, end_row = row_end - 1, start_column = 4, end_column = 4)
                ws.merge_cells(start_row = row, end_row = row_end - 1, start_column = 5, end_column = 5)

            if (row == row_end):
                row_end += 1

            row = row_end

    out_file_name = g_output_fld + "\\" + g_module_name + "_variables.xlsx"
    wb.save(out_file_name)



#/*****************************************************************************/
#/* 定義一覧出力                                                              */
#/*****************************************************************************/
def out_define_list():
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "typedef一覧"

    row = 2
    col = 2
    ws.cell(row, col    ).value = "ファイル名"
    ws.cell(row, col + 1).value = "type名"
    ws.cell(row, col + 2).value = "型"
    ws.cell(row, col + 3).value = "説明"

    row += 1
    for module in g_modules:
        for typedef in module.typedefs:
            ws.cell(row, col    ).value = module.name
            ws.cell(row, col + 1).value = typedef.name
            ws.cell(row, col + 2).value = typedef.type
            ws.cell(row, col + 3).value = typedef.summary
            row += 1


    ws = wb.create_sheet("enum定義一覧")
    row = 2
    col = 2
    ws.cell(row, col    ).value = "ファイル名"
    ws.cell(row, col + 1).value = "type名"
    ws.cell(row, col + 2).value = "型"
    ws.cell(row, col + 3).value = "説明"
    for module in g_modules:
        for typedef in module.typedefs:
            if (typedef.type == "enum"):
                ws.cell(row, col    ).value = module.name
                ws.cell(row, col + 1).value = typedef.name
                ws.cell(row, col + 2).value = typedef.type
                ws.cell(row, col + 3).value = typedef.summary
                row += 1
                for member in typedef.members:
                    ws.cell(row, col + 2).value = member.name
                    ws.cell(row, col + 3).value = member.type
                    ws.cell(row, col + 4).value = member.summary
                    row += 1


    ws = wb.create_sheet("構造体定義一覧")
    row = 2
    col = 2
    ws.cell(row, col    ).value = "ファイル名"
    ws.cell(row, col + 1).value = "type名"
    ws.cell(row, col + 2).value = "型"
    ws.cell(row, col + 3).value = "説明"
    for module in g_modules:
        for typedef in module.typedefs:
            if ((typedef.type == "struct") | (typedef.type == "union")):
                ws.cell(row, col    ).value = module.name
                ws.cell(row, col + 1).value = typedef.name
                ws.cell(row, col + 2).value = typedef.type
                ws.cell(row, col + 3).value = typedef.summary
                row += 1
                for member in typedef.members:
                    ws.cell(row, col + 2).value = member.type
                    ws.cell(row, col + 3).value = member.name
                    ws.cell(row, col + 4).value = member.summary
                    row += 1

    out_file_name = g_output_fld + "\\" + g_module_name + "_defines.xlsx"
    wb.save(out_file_name)



#/*****************************************************************************/
#/* 関数一覧出力                                                              */
#/*****************************************************************************/
def out_function_list():
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    ws.title = "関数一覧"

    row = 2
    col = 2
    ws.cell(row, col    ).value = "ファイル名"
    ws.cell(row, col + 1).value = "type名"
    ws.cell(row, col + 2).value = "型"
    ws.cell(row, col + 3).value = "説明"

    row += 1
    for module in g_modules:
        for function in module.functions:
            ws.cell(row, col    ).value = function.ret_type
            ws.cell(row, col + 1).value = function.name
            ws.cell(row, col + 2).value = function.name
            ws.cell(row, col + 3).value = function.name
            row += 1


    out_file_name = g_output_fld + "\\" + g_module_name + "_functions.xlsx"
    wb.save(out_file_name)


#/*****************************************************************************/
#/* Call Tree再帰処理                                                         */
#/*****************************************************************************/
def make_call_tree(is_called, level, function, pu_file, module, target_func):
    global g_tree_stack
    global g_tree_depend
    print("make_call_tree : %s" % function)

    #/* 再帰のチェック */
    for origin in g_tree_stack:
        if (origin == function):
            print("recursive call! : %s" % function)
            return;

    #/* まずスタックに関数名を積む */
    g_tree_stack.append(function)

    pu_text = "*" * level + "_ " + function
    print(pu_text, file= pu_file);

    if (is_called):
        refer = target_func.func_called
    else:
        refer = target_func.func_call

    for next_func_name in refer:
        print("make_call_tree next : %s" % next_func_name)
        next_func = module.get_function(next_func_name)
        if (next_func == None):
            for other_module in g_modules:
                if (other_module.name == module.name):
                    continue

                next_func = other_module.get_function(next_func_name)
                if (next_func != None):
                    break

        if (next_func == None):
            pu_text = "*" * (level + 1) + "_ " + next_func_name + "*"
            print(pu_text, file= pu_file);

            #/* 依存している外部関数の列挙 */
            dep = cDependency()
            dep.function = next_func_name
            for stack in g_tree_stack:
                dep.stack.append(stack)
            g_tree_depend.append(dep)
        else:
            make_call_tree(is_called, level + 1, next_func_name, pu_file, module, next_func)

    #/* スタックから自身を削除する */
    g_tree_stack.pop()
    return



#/*****************************************************************************/
#/* Call Tree生成                                                             */
#/*****************************************************************************/
def out_call_tree():
    global g_call_tree
    global g_called_tree
    global g_output_fld
    global g_tree_stack
    global g_tree_depend
    global g_jar_path
    global g_without_pu_convert
    print("out_call_tree")

    for target in g_call_tree:
        print("call tree : %s" % target)
        for module in g_modules:
            target_func = module.get_function(target)
            if (target_func != None):
                break

        if (target_func != None):
            print("call tree find : %s" % target)
            g_tree_depend = []
            file_path = g_output_fld + "\\" + target + "_call_tree.pu"
            pu_file = open(file_path, 'w', encoding="utf-8")
            print("@startmindmap", file= pu_file);
            make_call_tree(0, 1, target, pu_file, module, target_func)
            print("@endmindmap", file= pu_file);
            pu_file.close()

            file_path2 = g_output_fld + "\\" + target + "_call_tree_dependency.txt"
            dep_file = open(file_path2, 'w', encoding="utf-8")
            print("Dendency in %s call tree" % (target), file= dep_file);
            for dep in g_tree_depend:
                text = dep.function
                while (dep.stack):
                    text = text + "," + dep.stack.pop()
                print(text, file= dep_file);
            dep_file.close()
        else:
            print("call tree target not found : %s" % target)

        if (g_jar_path != ""):
            if (g_without_pu_convert == 0):
                cmd_text = "java -DPLANTUML_LIMIT_SIZE=16384 -jar " + g_jar_path + " " + file_path + " -charset UTF-8"
                print(cmd_text)
                subprocess.run(cmd_text, stdout=sys.stdout)

    for target in g_called_tree:
        print("called tree : %s" % target)
        for module in g_modules:
            target_func = module.get_function(target)
            if (target_func != None):
                break

        if (target_func != None):
            print("called tree find : %s" % target)
            g_tree_depend = []
            file_path = g_output_fld + "\\" + target + "_called_tree.pu"
            pu_file = open(file_path, 'w', encoding="utf-8")
            print("@startmindmap", file= pu_file);
            print("left side", file= pu_file);
            make_call_tree(1, 1, target, pu_file, module, target_func)
            print("@endmindmap", file= pu_file);
            pu_file.close()

            file_path2 = g_output_fld + "\\" + target + "called_tree_dependency.txt"
            dep_file = open(file_path2, 'w', encoding="utf-8")
            print("Dendency in %s call tree" % (target), file= dep_file);
            for dep in g_tree_depend:
                text = dep.function
                while (dep.stack):
                    text = text + "," + dep.stack.pop()
                print(text, file= dep_file);
            dep_file.close()
        else:
            print("called tree target not found : %s" % target)

        if (g_jar_path != ""):
            if (g_without_pu_convert == 0):
                cmd_text = "java -DPLANTUML_LIMIT_SIZE=16384 -jar " + g_jar_path + " " + file_path + " -charset UTF-8"
                print(cmd_text)
                subprocess.run(cmd_text, stdout=sys.stdout)





#/*****************************************************************************/
#/* メイン関数                                                                */
#/*****************************************************************************/
def main():
    check_command_line_option()
    make_directory(g_output_fld)
    read_setting_file()
    log_settings()

    if (g_exec_c_analyze):
        c_analyze()

    module_analyze()

    out_variable_list()
    out_define_list()
    out_function_list()
    out_call_tree()




if __name__ == "__main__":
    main()


